import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ITEMS = [
("it-001","Yogurt 1kg/Box","Greek Yogurt 1kg","1kg/Box",6,6,4),
("it-002","Yogurt 1kg/Box","Yuzu","1kg/Box",1,1,0),
("it-003","Yogurt 1kg/Box","Kyoho","1kg/Box",1,1,0),
("it-004","Yogurt 1kg/Box","Mint","1kg/Box",1,1,0),
("it-005","Yogurt 1kg/Box","Vanilla","1kg/Box",1,1,0),
("it-006","Yogurt 1kg/Box","Pineapple","1kg/Box",1,1,0),
("it-007","Yogurt 1kg/Box","Biscoff","1kg/Box",2,2,0),
("it-008","Yogurt 1kg/Box","Overnight oats biscoff","1kg/Box",1,1,0),
("it-009","Yogurt 1kg/Box","Plain Yogurt (ธรรมชาติ)","1kg/Box",6,6,4),
("it-010","Yogurt 500g/Box","Greek Yogurt 500g","500g/Box",12,8,7),
("it-011","Yogurt 500g/Box","Plain Yogurt 500g","500g/Box",4,2,2),
("it-012","ACAI","R-ACAI Bowl","Bowl",15,15,None),
("it-013","ACAI","S-ACAI Cup (mini)","Cup",15,15,None),
("it-014","Soft Serve / Ice Cream","น้ำ Ice cream / Soft Serve","2kg/Bag",4,None,None),
("it-015","Soft Serve / Ice Cream","Granola โรย Ice cream","g.",1,None,None),
("it-016","Soft Serve / Ice Cream","Softserve ถ้วยกระดาษ","ถ้วย",0,0,None),
("it-017","Shake แข็ง","Shake (แช่แข็ง)","ถ้วย",100,80,None),
("it-018","Drink / แยมกระปุก","Peanut Butter","กระปุก",None,None,None),
("it-019","Drink / แยมกระปุก","Water น้ำดื่ม","ขวด",12,12,None),
("it-020","Drink / แยมกระปุก","ถุงสตรอเบอรี่","ถุง",12,8,5),
("it-021","Drink / แยมกระปุก","ถุงบลูเบอรี่","ถุง",12,8,5),
("it-022","Drink / แยมกระปุก","ถุงธรรมชาติ","ถุง",12,10,5),
("it-023","Drink / แยมกระปุก","ถุงลิ้นจี่","ถุง",12,8,5),
("it-024","Drink / แยมกระปุก","ถุงยูส","ถุง",12,8,5),
("it-025","Drink / แยมกระปุก","ถุงพีช","ถุง",8,5,5),
("it-026","Cereals","Cornflakes Malt (M)","กระปุก",4,4,None),
("it-027","Cereals","Granola (M)","กระปุก",4,4,None),
("it-028","Cereals","Choc Chip Cookies","กระปุก",0,0,None),
("it-029","Toppings","Cookies Crumbs","750g/pack",1,1,0),
("it-030","Toppings","Oreo","454g/pack",2,2,0),
("it-031","Toppings","Choc Chips","300g/pack",1,1,0),
("it-032","Toppings","Cornflakes (Topping)","1,000g/box",1,1,0),
("it-033","Toppings","Granola (Topping)","2,000g/box",1,1,0),
("it-034","Toppings","Almond","400g/pack",1,1,None),
("it-035","Toppings","Pecan","400g/pack",1,1,None),
("it-036","Toppings","Walnut","400g/pack",1,1,None),
("it-037","Toppings","Coconut Chips","130g/box",0,0,None),
("it-038","Toppings","Chia Seed","400g/pack",1,1,0),
("it-039","Toppings","Flax Seed","400g/pack",1,1,0),
("it-040","Toppings","Cacao Nibs","400g/pack",1,1,0),
("it-041","Toppings","Grape Jelly","1,000g/pack",1,1,None),
("it-042","Toppings","Honey Jelly","1,000g/pack",1,1,None),
("it-043","Fruits","Strawberry (250g)","250g/box",4,4,0),
("it-044","Fruits","Strawberry (500g)","500g/box",None,None,None),
("it-045","Fruits","Blueberry (125g)","125g/box",None,None,None),
("it-046","Fruits","Blueberry (300g)","300g/box",3,3,0),
("it-047","Fruits","Blueberry (500g)","500g/box",None,None,None),
("it-048","Fruits","Apple Cinnamon","500g/pack",1,1,0),
("it-049","Fruits","Banana","ลูก",0,0,0),
("it-050","Sauces","Honey","1,000g/bottle",1,1,0),
("it-051","Sauces","Caramel","1,000g/bottle",1,1,0),
("it-052","Sauces","Peanut Butter Sauce","1,000g/pack",1,1,0),
("it-053","CUP/ถ้วย","Cup P (5oz)","50/pack",1,1,0),
("it-054","CUP/ถ้วย","Cup S (9oz)","50/pack",2,2,0),
("it-055","CUP/ถ้วย","Small Bowl","50/pack",1,1,0),
("it-056","CUP/ถ้วย","Cup (14oz)","50/pack",2,2,0),
("it-057","TOPPING CUP","Cup 3oz (Topping)","50/pack",2,2,0),
("it-058","TOPPING CUP","Cup+Lid (Big) 3oz","50/pack",2,2,0),
("it-059","TOPPING CUP","Cup+Lid (Small) 1oz","50/pack",2,2,0),
("it-060","LID/ฝา","Bowl Lid","50/pack",2,2,0),
("it-061","LID/ฝา","Smoothies Lid","50/pack",2,2,0),
("it-062","LID/ฝา","Lid S (92)","50/pack",2,2,0),
("it-063","LID/ฝา","Lid Top / Lid P (75)","50/pack",2,2,0),
("it-064","SPOON/ช้อน","Wood Spoon in bag","100/pack",1,1,0),
("it-065","SPOON/ช้อน","Wood Spoon","100/pack",3,3,0),
("it-066","SPOON/ช้อน","Tester Spoon ช้อนชิม","100/pack",1,1,0),
("it-067","SPOON/ช้อน","Short Spoon (ช้อนถ้วยP)","100/pack",2,2,0),
("it-068","SPOON/ช้อน","Straw/หลอด - ใหญ่","100/pack",2,2,0),
("it-069","SPOON/ช้อน","Straw/หลอด - เล็ก","100/pack",1,1,None),
("it-070","BAG/ถุง","Fail Bag ถุงฟลอย","ใบ",2,2,0),
("it-071","BAG/ถุง","ถุงกระดาษแก้วเดี่ยว","ใบ",40,40,0),
("it-072","BAG/ถุง","ถุงกระดาษแก้วคู่","ใบ",40,40,0),
("it-073","BAG/ถุง","ถุงกระดาษใหญ่","ใบ",30,30,0),
("it-074","BAG/ถุง","กระดาษกันหก","ห่อ (500ชิ้น)",1,1,None),
("it-075","BAG/ถุง","ฟอยล์แก้ว","อัน",30,30,0),
("it-076","BAG/ถุง","ฐานรองแก้วเดี่ยว","อัน",40,40,0),
("it-077","BAG/ถุง","ฐานรองแก้ว (คู่)","อัน",30,30,0),
("it-078","BAG/ถุง","Zip Bag Small / ถุงซิป","Pack",1,1,0),
("it-079","BAG/ถุง","Bag 4x14","Pack",2,2,None),
("it-080","BAG/ถุง","Bag 6x14","Pack",2,2,None),
("it-081","BAG/ถุง","Bag 8x14","Pack",2,2,None),
("it-082","BAG/ถุง","Bag 9x14","Pack",2,2,None),
("it-083","STICKER","Bag Sticker สติ๊กเกอร์ลิ้น","Sheet",4,4,0),
("it-084","STICKER","Sticker Roll (สก๊อตเทป)","Roll",2,2,None),
("it-085","ของใช้","Print Paper กระดาษปริ้น","Roll",3,3,0),
("it-086","ของใช้","Gloves YG S / ถุงมือ","Box",1,1,None),
("it-114","ของใช้","Gloves YG M / ถุงมือ","Box",1,1,None),
("it-115","ของใช้","Gloves YG L / ถุงมือ","Box",1,1,None),
("it-087","ของใช้","Black Bag 30x40 / ถุงขยะ","Roll",1,1,None),
("it-116","ของใช้","Black Bag 24x28 / ถุงขยะ","Roll",1,None,None),
("it-088","ของใช้","Dry Tissue / ทิชชู่แห้ง","Pack",3,3,None),
("it-089","ของใช้","Wet Tissue / ทิชชู่เปียก","Pack",2,2,None),
("it-090","ของใช้","Tissue ทิชชู่ลูกค้า","Pack",1,None,None),
("it-091","น้ำยาทำความสะอาด","น้ำยาถูพื้น","ขวด",1,1,None),
("it-092","น้ำยาทำความสะอาด","น้ำยาตัดไขมัน","ขวด",1,1,None),
("it-093","น้ำยาทำความสะอาด","น้ำยาล้างจาน","แพค",1,1,None),
("it-094","น้ำยาทำความสะอาด","น้ำยาอเนกประสงค์","ขวด",1,1,None),
("it-095","น้ำยาทำความสะอาด","น้ำยาล้างเครื่องไอดิม","ห่อ",1,1,None),
("it-096","Smoothies (Pre-packed)","Wake Up Call (1)","10ถุง/Pack",20,20,None),
("it-097","Smoothies (Pre-packed)","Energy Sip (2)","10ถุง/Pack",20,20,None),
("it-098","Smoothies (Pre-packed)","Yellow Madness (3)","10ถุง/Pack",20,20,None),
("it-099","Smoothies (Pre-packed)","Ready to Glow (5)","10ถุง/Pack",20,20,None),
("it-100","Yogurt Smoothies Powder","ผงโกโก้ (COCOA)","Bag",1,1,None),
("it-101","Yogurt Smoothies Powder","ผงมาคิ (MAQUI)","Bag",1,1,None),
("it-102","Yogurt Smoothies Powder","ผงคาม (CAMU)","Bag",1,1,None),
("it-103","Yogurt Smoothies Powder","น้ำเชื่อม (Syrup)","Bottle",1,1,None),
("it-104","Yogurt Shake Toppings","Biscoff Spread เล็ก","Bottle",2,2,0),
("it-105","Yogurt Shake Toppings","Biscoff Spread ใหญ่","Bottle",0,0,None),
("it-106","Yogurt Shake Toppings","ซอส Chocolate","500/Bag",1,1,0),
("it-107","Yogurt Shake Toppings","ซอส Strawberry","500/Bag",1,1,0),
("it-108","Yogurt Shake Toppings","ปีโป้","40 อัน/Bag",1,1,0),
("it-109","Yogurt Shake Toppings","ปีโป้ลิ้นจี่","กล่อง",1,1,0),
("it-110","Softserve Toppings","พิสตาชิโอ้เครป","กรัม",0,None,None),
("it-111","Softserve Toppings","พิสตาชิโอ้บัตเตอร์","กรัม",0,None,None),
("it-112","Softserve Toppings","พิสตาชิโอ้ท๊อปปิ้ง","กรัม",0,None,None),
("it-113","Cereals","Cranberry Cookies","กระปุก",0,0,None),
("it-117","น้ำยาทำความสะอาด","น้ำยาล้างผลไม้","ขวด",1,1,1),
("it-118","น้ำยาทำความสะอาด","น้ำยาซักผ้า","ขวด",1,None,None),
]

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="F2565C", end_color="F2565C", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
BASE_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()

# ── Sheet 1: Instructions ──────────────────────────────────────────
ws0 = wb.active
ws0.title = "Instructions"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 100
lines = [
    ("NCD Branch — Brand Tagging Workbook", True, 14),
    ("", False, 10),
    ("Purpose: NCD will sell both YC and Staple products in the same shift, using one shared", False, 11),
    ("stock/POS system. Each product needs a brand tag so sales and stock can be reported", False, 11),
    ("separately per brand even though staff work from one shared list.", False, 11),
    ("", False, 10),
    ("Sheet 'Existing Items' — 118 items currently in the system (all branches).", True, 11),
    ("  Column E 'Display Name at NCD' — ONLY fill this in if NCD staff need to see a", False, 11),
    ("    different name than what NVP/SND/KCN already show, to avoid confusion between", False, 11),
    ("    a YC version and a Staple version of the same kind of item (e.g. cups/lids/bags", False, 11),
    ("    printed with a brand logo — same product type, different physical item per brand).", False, 11),
    ("    Leave BLANK for everything else — the item keeps its normal name everywhere,", False, 11),
    ("    including at NCD. This never changes what NVP/SND/KCN staff see.", False, 11),
    ("  Column G 'Par NCD' — how many NCD should stock as normal full stock.", False, 11),
    ("    Leave BLANK if NCD will NOT carry this item at all (same rule as the other", False, 11),
    ("    branch columns — blank = not stocked there). This matches columns I/J/K which", False, 11),
    ("    show how NVP/SND/KCN already handle each item, for reference only — not editable.", False, 11),
    ("  Column H 'Brand for NCD' is pre-filled with 'yc' as a starting guess.", False, 11),
    ("    Only matters for items NCD actually carries (Par NCD is not blank).", False, 11),
    ("  Please review every row and change it where needed:", False, 11),
    ("    yc      = YC-only product, not sold under Staple", False, 11),
    ("    staple  = Staple-only product", False, 11),
    ("    shared  = same physical product sold under both brands with no branding difference", False, 11),
    ("  Click any cell in column H to pick from the dropdown.", False, 11),
    ("", False, 10),
    ("Note: showing a different name only at NCD (column E) needs a small feature that does", False, 11),
    ("not exist in the app yet — NCD is not even set up as a branch in the system yet", False, 11),
    ("(opens Sep 2569). This is captured here now so it's ready to build when NCD launches.", False, 11),
    ("", False, 10),
    ("Sheet 'New Staple Items' — add any product that does not exist in the system yet.", True, 11),
    ("  Despite the sheet name, this also covers brand-new items that will be SHARED between", False, 11),
    ("  YC and Staple at NCD but don't exist anywhere else yet (e.g. a new bag design used by", False, 11),
    ("  both brands at NCD only, different from the bag design NVP/SND/KCN currently use).", False, 11),
    ("  Set the Brand column to yc / staple / shared accordingly.", False, 11),
    ("  A few rows are already pre-filled based on what's been confirmed so far — check them,", False, 11),
    ("  then add your own rows below.", False, 11),
    ("  Category can be an existing one (see 'Existing Items' sheet) or a brand-new category name.", False, 11),
    ("  Par level = normal full-stock quantity to keep on the shelf at NCD.", False, 11),
    ("", False, 10),
    ("When done, send this file back and it will be imported directly into the system.", True, 11),
]
r = 1
for text, bold, size in lines:
    c = ws0.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT_NAME, bold=bold, size=size, color="542916" if bold else "2B241C")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1
ws0.row_dimensions[1].height = 22

# ── Sheet 2: Existing Items ─────────────────────────────────────────
# แพรทัก 2026-08-05 — ถ้วย 14oz พิมพ์โลโก้คนละแบบ YC/Staple แยกกัน ต้องแยกนับสต็อก
# ต้องการให้พนักงาน NCD เห็นชื่อระบุแบรนด์ชัดๆ กันสับสน (สาขาอื่นเห็นชื่อเดิม)
# Display Name at NCD แปลเป็นอังกฤษเลย เพราะคอลัมน์นี้คือชื่อที่ NCD เห็นจริง (พนักงาน NCD เป็นต่างชาติ)
#
# ⚠️ ถุงกระดาษแก้วเดี่ยว/คู่/ใหญ่ (it-071/072/073) แพรแก้อีกรอบ 2026-08-05 — ลายที่ NCD ใช้
# เป็นคนละลายกับที่ NVP/SND/KCN ใช้อยู่ปัจจุบัน (แม้จะ shared ระหว่าง YC/Staple ที่ NCD เอง)
# แปลว่า it-071/072/073 เดิม "ไม่ใช่" ของที่ NCD ใช้เลย — NCD ต้องมีของใหม่ 3 อัน (ดูชีต
# New Staple Items แทน) item เดิมพวกนี้ NCD ไม่ carry (Par NCD ว่าง) เหมือนตอนที่ยังไม่ได้แตะเลย
NCD_DISPLAY_NAME = {
    "it-056": "YC Cup (14oz)",
}
NCD_BRAND_SUGGESTION: dict = {}

ws1 = wb.create_sheet("Existing Items")
headers = ["No.", "Item ID", "Category", "Item Name", "Display Name at NCD (blank = same)", "Unit",
           "Par NCD (blank = not carried at NCD)", "Brand for NCD (yc / staple / shared)",
           "Par NVP (ref)", "Par SND (ref)", "Par KCN (ref)"]
widths = [5, 10, 22, 34, 26, 14, 22, 30, 12, 12, 12]
for col, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[1].height = 30
ws1.freeze_panes = "A2"

dv = DataValidation(type="list", formula1='"yc,staple,shared"', allow_blank=False, showDropDown=False)
dv.error = "Choose yc, staple, or shared from the list."
dv.errorTitle = "Invalid brand"
ws1.add_data_validation(dv)

# คอลัมน์อ้างอิง (ยกมาจาก NVP/SND/KCN) ใช้บอกว่าสาขาอื่นๆ ขายไหม — บอกใบ้เฉยๆ ไม่ได้ให้แก้
for i, (item_id, category, name, unit, par_nvp, par_snd, par_kcn) in enumerate(ITEMS, start=1):
    row = i + 1
    display_name = NCD_DISPLAY_NAME.get(item_id)
    brand_suggestion = NCD_BRAND_SUGGESTION.get(item_id, "yc")
    values = [i, item_id, category, name, display_name, unit, None, brand_suggestion, par_nvp, par_snd, par_kcn]
    for col, val in enumerate(values, start=1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = BASE_FONT
        cell.border = BORDER
        if col in (1, 7, 9, 10, 11):
            cell.alignment = CENTER
        elif col == 8:
            cell.alignment = CENTER
            cell.fill = INPUT_FILL
        else:
            cell.alignment = WRAP_LEFT
        if col in (5, 7):
            cell.fill = INPUT_FILL
    dv.add(ws1.cell(row=row, column=8))

ws1["E1"].comment = Comment("Only fill in if NCD needs a different, clearer name than the one shown at other branches (e.g. to tell a YC-branded item apart from a Staple-branded version of the same kind of thing). Leave blank otherwise — this never changes what other branches see.", "System")
ws1["G1"].comment = Comment("Fill in a number if NCD carries this item (normal full-stock qty). Leave blank if NCD does NOT carry it at all.", "System")
ws1["H1"].comment = Comment("Only matters if Par NCD (column G) is filled in. Pick yc / staple / shared from the dropdown.", "System")

# ── Sheet 3: New Staple Items ───────────────────────────────────────
# ชื่อชีตยังคงเดิม แต่ใช้ใส่ item ใหม่ทุกชนิดที่ NCD ต้องมี ไม่ใช่แค่ของ staple ล้วน — บาง
# item (เช่นถุงกระดาษลายใหม่) เป็น shared ระหว่าง YC/Staple ที่ NCD เอง จึงมีคอลัมน์ Brand ให้เลือก
ws2 = wb.create_sheet("New Staple Items")
headers2 = ["No.", "Item Name (English)", "Category", "Unit", "Brand (yc / staple / shared)", "Par NCD", "Notes"]
widths2 = [5, 34, 22, 16, 22, 12, 40]
for col, (h, w) in enumerate(zip(headers2, widths2), start=1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[1].height = 26
ws2.freeze_panes = "A2"

dv2 = DataValidation(type="list", formula1='"yc,staple,shared"', allow_blank=False, showDropDown=False)
dv2.error = "Choose yc, staple, or shared from the list."
dv2.errorTitle = "Invalid brand"
ws2.add_data_validation(dv2)

example = [1, "Belgian Waffle Mix", "Staple Bakery", "1kg/Bag", "staple", 4, "Example row — delete before sending back, or add rows below"]
for col, val in enumerate(example, start=1):
    cell = ws2.cell(row=2, column=col, value=val)
    cell.font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
    cell.border = BORDER
    cell.alignment = CENTER if col in (1, 6) else WRAP_LEFT

# แพรทัก 2026-08-05 — ถ้วย 14oz: YC/Staple พิมพ์คนละลาย แยกเป็นคนละ item กันเลย (brand yc/staple)
# ถุงกระดาษแก้วเดี่ยว/คู่/ใหญ่: ลายใหม่ที่ NCD ใช้เป็นคนละลายกับที่ NVP/SND/KCN ใช้อยู่ (it-071/072/073)
# แต่ YC กับ Staple ที่ NCD ใช้ลายใหม่นี้ร่วมกัน (brand shared) — คนละเรื่องกับถ้วยที่แยกกันจริงๆ
KNOWN_NEW_ROWS = [
    ("Staple Cup (14oz)", "CUP/ถ้วย", "50/pack", "staple", None,
     "Same size/spec as YC Cup (14oz, it-056) but printed with Staple logo — separate stock item, not shared. Fill in Par NCD."),
    ("NCD Single Cup Paper Bag (new design)", "Bags", "ใบ", "shared", None,
     "New bag design used ONLY at NCD, shared by both YC and Staple. NOT the same item as YC's existing ถุงกระดาษแก้วเดี่ยว (it-071) used at NVP/SND/KCN — that one stays a separate, unrelated item. Fill in Par NCD."),
    ("NCD Double Cup Paper Bag (new design)", "Bags", "ใบ", "shared", None,
     "New bag design used ONLY at NCD, shared by both YC and Staple. NOT the same item as YC's existing ถุงกระดาษแก้วคู่ (it-072) used at NVP/SND/KCN — that one stays a separate, unrelated item. Fill in Par NCD."),
    ("NCD Large Paper Bag (new design)", "Bags", "ใบ", "shared", None,
     "New bag design used ONLY at NCD, shared by both YC and Staple. NOT the same item as YC's existing ถุงกระดาษใหญ่ (it-073) used at NVP/SND/KCN — that one stays a separate, unrelated item. Fill in Par NCD."),
]

for row in range(3, 43):
    known_idx = row - 3
    known = KNOWN_NEW_ROWS[known_idx] if known_idx < len(KNOWN_NEW_ROWS) else None
    row_values = [row - 1] + (list(known) if known else [None, None, None, None, None, None])
    for col, val in enumerate(row_values, start=1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.fill = INPUT_FILL
        cell.alignment = CENTER if col in (1, 6) else WRAP_LEFT
    dv2.add(ws2.cell(row=row, column=5))

import os
out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "NCD_Brand_Tagging.xlsx")
wb.save(out_path)
print("Saved:", out_path)
print("Existing items rows:", len(ITEMS))
