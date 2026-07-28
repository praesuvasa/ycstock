"""
banthuek.py — บันทึก [DD/MM/YYYY] [BRANCH]
Usage: python3 banthuek.py 02/06/2026 NVP
"""
import sys
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Protection

WB_PATH = '/sessions/cool-pensive-turing/mnt/YC Inventory Analysis/YC_Stock_Tracker.xlsx'

# ─── Input col → DB col letter ───────────────────────────
INPUT_TO_DB = {6:'F', 7:'G', 8:'H', 9:'I', 10:'J', 11:'K', 12:'L'}

# Packaging rows: E = plain input → restored as carry-forward formula after บันทึก
PLAIN_E_ROWS = {75, 76, 77, 79, 80, 81}

FILLS = {
    6:  PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    7:  PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    8:  PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid'),
    9:  PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),
    10: PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),
    11: PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
    12: PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
}
FILL_E      = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
FILL_LOCKED = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
FONT_FORMULA = Font(name='Calibri', size=11, italic=True, color='555555')
FONT_BLUE    = Font(name='Calibri', size=11, italic=True, color='1565C0')
FONT_LOCKED  = Font(name='Calibri', size=11, italic=True, color='757575')
CENTER = Alignment(horizontal='center', vertical='center')

# Specific cells to lock (gray fill + locked=True) — re-applied after each restore
def _build_locked_cells():
    s = set()
    def add(col, rows):
        for r in (rows if hasattr(rows, '__iter__') else [rows]):
            s.add((col, r))
    # J column (col=10)
    add(10, range(13, 17))    # J13-16
    add(10, range(19, 32))    # J19-31
    add(10, range(61, 75))    # J61-74
    add(10, range(82, 87))    # J82-86
    add(10, range(90, 100))   # J90-99
    # I column (col=9)
    add(9,  [52])             # I52
    add(9,  range(75, 82))    # I75-81
    add(9,  range(87, 90))    # I87-89
    add(9,  range(100, 104))  # I100-103
    # G column (col=7)
    add(7,  [11])             # G11
    add(7,  range(13, 17))    # G13-16
    add(7,  range(19, 52))    # G19-51
    add(7,  range(53, 56))    # G53-55
    add(7,  range(56, 75))    # G56-74
    add(7,  [78])             # G78
    add(7,  range(82, 87))    # G82-86
    add(7,  range(90, 100))   # G90-99
    add(7,  range(104, 114))  # G104-113
    # F column (col=6)
    add(6,  [52])             # F52
    add(6,  range(75, 78))    # F75-77
    add(6,  range(79, 82))    # F79-81
    add(6,  range(87, 90))    # F87-89
    add(6,  range(100, 104))  # F100-103
    return s

LOCKED_CELLS = _build_locked_cells()


def apply_locked_cells(ws):
    """Apply gray fill + locked=True to specified cells (called after restore_fl_formulas)."""
    for (col, row) in LOCKED_CELLS:
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_LOCKED
        cell.font = FONT_LOCKED
        cell.protection = Protection(locked=True)

def fl_formula(r, db_col):
    """Read from DB for today's date; blank if DB cell is blank, "" if no DB row."""
    s = 122
    m = f'บันทึกสต็อครายวัน!$M${s}:$M$9999'
    key = f'TEXT($B$2,"YYYYMMDD")&"|"&$E$2&"|"&B{r}'
    match = f'MATCH({key},{m},0)'
    idx = f'INDEX(บันทึกสต็อครายวัน!${db_col}${s}:${db_col}$9999,{match})'
    return f'=IFERROR(IF(ISBLANK({idx}),"",{idx}),"")'

def i_hybrid_formula(r):
    """I = คงเหลือ ชิ้น: DB value if saved, else D+F-H (treats ""→0, no negatives)."""
    s = 122
    fl = (f'INDEX(บันทึกสต็อครายวัน!$I${s}:$I$9999,'
          f'MATCH(TEXT($B$2,"YYYYMMDD")&"|"&$E$2&"|"&B{r},'
          f'บันทึกสต็อครายวัน!$M${s}:$M$9999,0))')
    comp = (f'IFERROR(IF(D{r}="",0,D{r})+IF(F{r}="",0,F{r})-IF(H{r}="",0,H{r}),"")')
    return f'=IFERROR(IF({fl}="",{comp},{fl}),{comp})'

def j_hybrid_formula(r):
    """J = คงเหลือ g: DB value if saved, else MAX(E+G-H, 0) (treats ""→0, no negatives)."""
    s = 122
    fl = (f'INDEX(บันทึกสต็อครายวัน!$J${s}:$J$9999,'
          f'MATCH(TEXT($B$2,"YYYYMMDD")&"|"&$E$2&"|"&B{r},'
          f'บันทึกสต็อครายวัน!$M${s}:$M$9999,0))')
    comp = (f'IFERROR(MAX(IF(E{r}="",0,E{r})+IF(G{r}="",0,G{r})-IF(H{r}="",0,H{r}),0),"")')
    return f'=IFERROR(IF({fl}="",{comp},{fl}),{comp})'

def d_carryforward_formula(r):
    """D for all item rows: today's DB col D if saved; fallback = most recent I (chronological max date).
    Uses SUMPRODUCT(MAX(...)) to find the max date serial across non-chronological DB rows.
    This avoids both LOOKUP (position-order) and MAXIFS (LO date-type comparison) issues."""
    s = 122
    a  = f'บันทึกสต็อครายวัน!$A${s}:$A$9999'
    bc = f'บันทึกสต็อครายวัน!$B${s}:$B$9999'
    cc = f'บันทึกสต็อครายวัน!$C${s}:$C$9999'
    m  = f'บันทึกสต็อครายวัน!$M${s}:$M$9999'
    today_key = f'TEXT($B$2,"YYYYMMDD")&"|"&$E$2&"|"&B{r}'
    today_D   = f'INDEX(บันทึกสต็อครายวัน!$D${s}:$D$9999,MATCH({today_key},{m},0))'
    # SUMPRODUCT(MAX(...)) forces array eval: max date serial where A<B2, B=branch, C=item
    max_date  = f'SUMPRODUCT(MAX(({a}<$B$2)*({bc}=$E$2)*({cc}=B{r})*{a}))'
    mr_key    = f'TEXT({max_date},"YYYYMMDD")&"|"&$E$2&"|"&B{r}'
    mr_I      = f'IFERROR(IF({max_date}=0,"",INDEX(บันทึกสต็อครายวัน!$I${s}:$I$9999,MATCH({mr_key},{m},0))),"")'
    return f'=IFERROR(IF({today_D}="",{mr_I},{today_D}),{mr_I})'

def e_carryforward_formula(r):
    """E for all rows: today's DB col E if saved; fallback = most recent J (chronological max date).
    Uses SUMPRODUCT(MAX(...)) for correct cross-sheet array evaluation in LibreOffice."""
    s = 122
    a  = f'บันทึกสต็อครายวัน!$A${s}:$A$9999'
    bc = f'บันทึกสต็อครายวัน!$B${s}:$B$9999'
    cc = f'บันทึกสต็อครายวัน!$C${s}:$C$9999'
    m  = f'บันทึกสต็อครายวัน!$M${s}:$M$9999'
    today_key = f'TEXT($B$2,"YYYYMMDD")&"|"&$E$2&"|"&B{r}'
    today_E   = f'INDEX(บันทึกสต็อครายวัน!$E${s}:$E$9999,MATCH({today_key},{m},0))'
    max_date  = f'SUMPRODUCT(MAX(({a}<$B$2)*({bc}=$E$2)*({cc}=B{r})*{a}))'
    mr_key    = f'TEXT({max_date},"YYYYMMDD")&"|"&$E$2&"|"&B{r}'
    mr_J      = f'IFERROR(IF({max_date}=0,"",INDEX(บันทึกสต็อครายวัน!$J${s}:$J$9999,MATCH({mr_key},{m},0))),"")'
    return f'=IFERROR(IF({today_E}="",{mr_J},{today_E}),{mr_J})'

def restore_fl_formulas(ws):
    """Restore INDEX-MATCH formulas in D-L for all item rows,
       E carry-forward for packaging rows, and I/J lock."""
    FILL_D = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    for r in range(4, 118):
        item = ws.cell(row=r, column=2).value
        if not item or str(item).startswith('🔢'):
            continue
        # Restore D (col 4) — ยกมา (packs): today's DB D if saved, else most recent I
        cell = ws.cell(row=r, column=4)
        cell.value = d_carryforward_formula(r)
        cell.fill = FILL_D
        cell.font = FONT_FORMULA
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)
        # Restore E (col 5) — all rows: today's DB E if saved, else most recent J
        cell = ws.cell(row=r, column=5)
        cell.value = e_carryforward_formula(r)
        cell.fill = FILL_E
        cell.font = FONT_BLUE if r in PLAIN_E_ROWS else FONT_FORMULA
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)
        # Restore F-L: all columns use fl_formula (blank if no DB entry for today)
        for input_col, db_col in INPUT_TO_DB.items():
            cell = ws.cell(row=r, column=input_col)
            cell.value = fl_formula(r, db_col)
            cell.alignment = CENTER
            cell.fill = FILLS[input_col]
            cell.font = FONT_FORMULA
            cell.protection = Protection(locked=False)

def get_carryforward(ws_db, branch, item, before_date_str):
    """Return (most_recent_I, most_recent_J) for branch/item from rows before before_date_str."""
    best_i_date = best_j_date = ''
    best_i = best_j = 0
    for r in range(121, 9999):
        a = ws_db.cell(row=r, column=1).value
        if a is None:
            break
        b = ws_db.cell(row=r, column=2).value
        c = ws_db.cell(row=r, column=3).value
        if b != branch or c != item:
            continue
        ds = a.strftime('%Y%m%d') if isinstance(a, datetime) else str(a)[:8].replace('-', '')
        if ds >= before_date_str:
            continue
        i_val = ws_db.cell(row=r, column=9).value
        j_val = ws_db.cell(row=r, column=10).value
        if i_val is not None and ds > best_i_date:
            best_i_date = ds
            best_i = i_val
        if j_val is not None and ds > best_j_date:
            best_j_date = ds
            best_j = j_val
    return best_i, best_j


def run(date_str, branch):
    save_date = datetime.strptime(date_str, '%d/%m/%Y')
    target_key_prefix = save_date.strftime('%Y%m%d') + '|' + branch + '|'
    target_date_str = save_date.strftime('%Y%m%d')

    # ── 1. Read input sheet (plain values only — no formula cache dependency) ──
    # Use non-data_only so we can distinguish formulas (str starting with "=") from user-entered values.
    # Formula cells are treated as None; only plain numbers the user typed are used.
    wb_in = openpyxl.load_workbook(WB_PATH)
    ws_in = wb_in['📝 กรอกข้อมูล']

    def plain_val(cell_val):
        """Return value if it's a plain number/string; None if it's a formula."""
        if isinstance(cell_val, str) and cell_val.startswith('='):
            return None
        return cell_val

    rows_data = []
    for r in range(4, 118):
        item = ws_in.cell(row=r, column=2).value
        if not item or str(item).startswith('🔢'):
            continue
        # D-L (cols 4-12), filter out formulas
        vals = [plain_val(ws_in.cell(row=r, column=c).value) for c in range(4, 13)]
        # Include row if user typed a plain value in F, G, H, I, or J (indices 2,3,4,5,6)
        if any(vals[i] is not None for i in [2, 3, 4, 5, 6]):
            rows_data.append((r, item, *vals))   # (row, item, D, E, F, G, H, I, J, K, L)

    print(f'Input rows: {len(rows_data)}')

    # ── 2. Load main workbook for editing ──
    wb = openpyxl.load_workbook(WB_PATH)
    ws_main = wb['บันทึกสต็อครายวัน']
    ws_input = wb['📝 กรอกข้อมูล']

    # Build KEY → row map
    key_to_row = {}
    next_empty = None
    for r in range(121, 9999):
        a = ws_main.cell(row=r, column=1).value
        if a is None:
            next_empty = r
            break
        b = ws_main.cell(row=r, column=2).value
        c = ws_main.cell(row=r, column=3).value
        ds = a.strftime('%Y%m%d') if isinstance(a, datetime) else str(a)[:8].replace('-','')
        key_to_row[f'{ds}|{b}|{c}'] = r

    print(f'DB rows: {len(key_to_row)}, next empty: {next_empty}')

    # ── 3. Upsert ──
    updated = inserted = 0
    for row_data in rows_data:
        r_in, item, d_input, e_input, f, g, h, i_raw, j_raw, k, l = row_data
        key = target_key_prefix + item

        f_n = f if isinstance(f, (int, float)) else 0
        g_n = g if isinstance(g, (int, float)) else 0
        h_n = h if isinstance(h, (int, float)) else 0

        if key in key_to_row:
            # UPDATE: use existing DB D/E to preserve manual corrections; update F,G,H; recompute I
            r = key_to_row[key]
            d_db = ws_main.cell(row=r, column=4).value
            e_db = ws_main.cell(row=r, column=5).value
            d_n = d_db if isinstance(d_db, (int, float)) else 0
            # I: use user-typed value if entered; else compute D+F-H
            if isinstance(i_raw, (int, float)):
                i_save = int(i_raw) if i_raw == int(i_raw) else i_raw
            else:
                i_save = max(d_n + f_n - h_n, 0)
            existing_j = ws_main.cell(row=r, column=10).value
            j_save = j_raw if isinstance(j_raw, (int, float)) else existing_j
            for ci, v in enumerate([d_db, e_db, f, g, h, i_save, j_save, k, l], start=4):
                ws_main.cell(row=r, column=ci).value = v
            updated += 1
        else:
            # INSERT: carry D/E from most recent DB row before this date
            d_cf, e_cf = get_carryforward(ws_main, branch, item, target_date_str)
            d_use = d_input if isinstance(d_input, (int, float)) else d_cf
            e_use = e_input if isinstance(e_input, (int, float)) else e_cf
            # I: use user-typed value if entered; else compute D+F-H
            if isinstance(i_raw, (int, float)):
                i_save = int(i_raw) if i_raw == int(i_raw) else i_raw
            else:
                i_save = max((d_use if isinstance(d_use, (int, float)) else 0) + f_n - h_n, 0)
            r = next_empty
            date_cell = ws_main.cell(row=r, column=1)
            date_cell.value = save_date
            date_cell.number_format = 'd/m/yyyy'
            ws_main.cell(row=r, column=2).value = branch
            ws_main.cell(row=r, column=3).value = item
            j_save = j_raw if isinstance(j_raw, (int, float)) else None
            for ci, v in enumerate([d_use, e_use, f, g, h, i_save, j_save, k, l], start=4):
                ws_main.cell(row=r, column=ci).value = v
            ws_main.cell(row=r, column=13).value = f'{target_date_str}|{branch}|{item}'
            next_empty += 1
            inserted += 1

    print(f'Updated: {updated}, Inserted: {inserted}')

    # ── 4. Restore F-L INDEX-MATCH formulas ──
    restore_fl_formulas(ws_input)
    apply_locked_cells(ws_input)
    print('F-L formulas restored ✓')

    # ── 4b. Advance B2 to next day so LO re-save caches D/E with correct carryforward ──
    from datetime import timedelta
    next_day = save_date + timedelta(days=1)
    b2_cell = ws_input['B2']
    b2_cell.value = next_day
    b2_cell.number_format = 'd/m/yyyy'

    wb.save(WB_PATH)
    print(f'✓ บันทึก {date_str} {branch} เรียบร้อย — DB row สุดท้าย: {next_empty-1}')
    print(f'  B2 → {next_day.strftime("%d/%m/%Y")} (พร้อมกรอกวันถัดไป)')

    # ── 5. Lock styles already applied inside restore_fl_formulas above ──
    import subprocess, os

    # ── 6. Re-save via LibreOffice to ensure Excel-compatible xlsx ──
    # openpyxl stores strings as inlineStr (no sharedStrings.xml); Excel Mac may reject this.
    # LO re-save produces a standard xlsx with sharedStrings and proper structure.
    lo = '/usr/bin/soffice'
    if os.path.exists(lo):
        import tempfile, shutil
        with tempfile.TemporaryDirectory() as tmpdir:
            subprocess.run(
                [lo, '--headless', '--convert-to', 'xlsx', WB_PATH, '--outdir', tmpdir],
                check=True, capture_output=True
            )
            lo_out = os.path.join(tmpdir, os.path.basename(WB_PATH))
            if os.path.exists(lo_out):
                shutil.copy2(lo_out, WB_PATH)
                print('LibreOffice re-save ✓')
            else:
                print('LibreOffice re-save skipped (output not found)')
    # Remove stale Excel lock file if present
    lock = os.path.join(os.path.dirname(WB_PATH), '~$' + os.path.basename(WB_PATH))
    if os.path.exists(lock):
        try:
            os.remove(lock)
            print('Stale lock file removed ✓')
        except Exception:
            pass

    # ── 7. Repair sales sheet view zone formulas ──────────────────────────────
    # LO re-save strips SUMIFS formulas in บันทึกยอดขาย when they return numeric
    # values (LO caches result and removes <f> element). Re-inject them via XML.
    repair_sales_view_formulas(WB_PATH)
    print('Sales view zone formulas repaired ✓')


def repair_sales_view_formulas(wb_path):
    """
    Re-apply SUMIFS formulas to view zone C/D/G/H cells (rows 3-34) in sheet3
    (บันทึกยอดขาย) where LO re-save stripped them to plain numeric values.
    Uses direct zipfile XML manipulation — no openpyxl, no LO needed.
    """
    import zipfile, re as _re, shutil as _shutil

    S_VIEW = '178'

    def _sumifs_cell(col, r):
        col_u = col.upper()
        f = (
            f'IFERROR(IF($A{r}=&quot;&quot;,&quot;&quot;,'
            f'IF(SUMIFS(${col_u}$41:${col_u}$500,$A$41:$A$500,$A{r},$B$41:$B$500,$I$1)=0,&quot;&quot;,'
            f'SUMIFS(${col_u}$41:${col_u}$500,$A$41:$A$500,$A{r},$B$41:$B$500,$I$1))),&quot;&quot;)'
        )
        return f'<c r="{col_u}{r}" s="{S_VIEW}" t="str"><f aca="false">{f}</f><v></v></c>'

    with zipfile.ZipFile(wb_path) as z:
        names = z.namelist()
        # Identify sales sheet (sheet3)
        raw = {n: z.read(n) for n in names}

    s3 = raw['xl/worksheets/sheet3.xml'].decode('utf-8')
    changed = 0

    # Build set of date serials that exist in entry zone (rows 41+)
    # Only replace view zone plain values for dates that have entry zone data
    entry_dates = set()
    for entry_row in _re.finditer(r'<row r="(\d+)"[^>]*>.*?</row>', s3, _re.DOTALL):
        rnum = int(entry_row.group(1))
        if rnum < 41: continue
        # Check A cell has a date value (t="n" or no t attr)
        a_m = _re.search(r'<c r="A%d"[^>]*>.*?<v>([^<]+)</v>' % rnum, entry_row.group(), _re.DOTALL)
        b_m = _re.search(r'<c r="B%d"[^>]*t="s"[^>]*><v>[^<]+</v>' % rnum, entry_row.group())
        if a_m and b_m:
            entry_dates.add(a_m.group(1))  # date serial as string

    for r in range(3, 35):
        # Get this view zone row's date serial from A column
        row_m = _re.search(r'<row r="%d"[^>]*>.*?</row>' % r, s3, _re.DOTALL)
        if not row_m: continue
        a_m = _re.search(r'<c r="A%d"[^>]*>.*?<v>([^<]+)</v>' % r, row_m.group(), _re.DOTALL)
        if not a_m: continue
        row_date = a_m.group(1)
        # Only repair if entry zone has data for this date (otherwise plain values = user-typed)
        if row_date not in entry_dates:
            continue
        for col in ['C', 'D', 'G', 'H']:
            ref = f'{col}{r}'
            # Match plain numeric cell (no formula)
            plain_pat = r'<c r="%s" s="[^"]*" t="n"><v>[^<]*</v></c>' % ref
            new_cell = _sumifs_cell(col, r)
            new_s3 = _re.sub(plain_pat, new_cell, s3)
            if new_s3 != s3:
                s3 = new_s3
                changed += 1

    if changed:
        raw['xl/worksheets/sheet3.xml'] = s3.encode('utf-8')
        tmp = wb_path + '.repair_tmp'
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as z:
            for name, data in raw.items():
                z.writestr(name, data)
        _shutil.move(tmp, wb_path)
        print(f'  Repaired {changed} view zone formula cell(s) ✓')
    else:
        print('  No view zone formula cells needed repair')


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python3 banthuek.py DD/MM/YYYY BRANCH')
        sys.exit(1)
    run(sys.argv[1], sys.argv[2])
